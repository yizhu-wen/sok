"""
Build a full-results Excel workbook for a benchmark dataset.

Examples:
  python3 scripts/18_dataset_full_excel.py \
      --dataset speech_ljspeech \
      --out results/speech_ljspeech_full.xlsx \
      --suffixes __plain__

  python3 scripts/18_dataset_full_excel.py \
      --dataset music_moisesdb \
      --out results/music_moisesdb_full.xlsx \
      --suffixes __plain__,recovery
"""
import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_DIR = Path(__file__).resolve().parent.parent
ALGORITHMS = [
    "audioseal",
    "wavmark",
    "silentcipher",
    "timbre",
    "dnn_watermark",
    "audiowmark",
    "aware",
    "fsvc",
    "patchwork",
    "normspace",
]

ALGO_LABELS = {
    "audioseal": "ICML'24 / AudioSeal",
    "wavmark": "ICLR'24 / WavMark",
    "silentcipher": "Interspeech'24 / SilentCipher",
    "timbre": "NDSS'24 / Timbre",
    "dnn_watermark": "DSP'22 / DNN-WM",
    "audiowmark": "audiowmark",
    "aware": "arXiv'25 / AWARE",
    "fsvc": "TASLP'21 / FSVC",
    "patchwork": "TASLP'17 / Patchwork",
    "normspace": "EURASIP'19 / NormSpace",
}

SECTIONS = [
    ("No Distortion", "no_distortion", [None]),
    ("Pitch Shift (cents)", "pitch_shift", [-100, -50, -25, -12.5, -6.25, 6.25, 12.5, 25, 50, 100]),
    ("Time Stretch", "time_stretch", [0.25, 0.5, 0.75, 0.9, 0.95, 1.05, 1.1, 1.25, 1.5, 1.75, 2.0]),
    ("Gaussian Noise (SNR)", "gaussian_noise", [-20, -15, -10, -5, 0, 5, 10, 20, 30, 40]),
    ("Bitcrush (bits)", "bitcrush", [2, 4, 6, 8, 10, 12]),
    ("MP3 Compression", "mp3_compression", [8, 16, 24, 32, 40, 48, 56, 64, 128, 192, 256, 320]),
    ("Background Noise", "background_noise", [
        "DKITCHEN", "DLIVING", "DWASHING",
        "NFIELD", "NPARK", "NRIVER",
        "OHALLWAY", "OMEETING", "OOFFICE",
        "PCAFETER", "PRESTO", "PSTATION",
        "SCAFE", "SPSQUARE", "STRAFFIC",
        "TBUS", "TCAR", "TMETRO",
    ]),
    ("Cutting Audio", "cutting_audio", [25, 50, 75, 90, 95, 97]),
    ("High-Pass Filter (Hz)", "high_pass_filter", [500, 1000, 1500, 2000, 2500, 3000, 3500, 4000]),
    ("Low-Pass Filter (Hz)", "low_pass_filter", [500, 1000, 1500, 2000, 2500, 3000, 3500, 4000]),
    ("Sample Suppression", "sample_suppression", [0.5, 1, 2.5, 5, 10, 25]),
    ("Resampling (kHz)", "resampling", [2, 4, 8, 16, 22.05]),
    ("Reverberation", "reverberation", [
        "bin_aula_carolina", "bin_booth", "bin_lecture", "bin_meeting",
        "bin_office", "bin_stairway",
        "phone_bathroom", "phone_corridor", "phone_kitchen", "phone_lecture",
        "phone_lecture1", "phone_meeting", "phone_office",
        "phone_stairway", "phone_stairway1", "phone_stairway2",
    ]),
]

HDR_FILL = PatternFill("solid", fgColor="2E4057")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
MISS_FILL = PatternFill("solid", fgColor="FCE4D6")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
SECTION_COLORS = [
    "4472C4", "70AD47", "ED7D31", "9E480E", "A5A5A5",
    "FFC000", "5B9BD5", "C00000", "00B0F0", "7030A0",
    "FF0000", "92D050", "00B050", "0070C0", "002060", "833C00",
]


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument(
        "--suffixes",
        required=True,
        help="Comma-separated suffix preference list. Use __plain__ for algo.json.",
    )
    parser.add_argument(
        "--algorithms",
        default="",
        help="Optional comma-separated algorithm list. Defaults to all known algorithms.",
    )
    return parser.parse_args()


def _result_path(results_dir, algo, suffix):
    if suffix == "__plain__":
        return results_dir / f"{algo}.json"
    return results_dir / f"{algo}_{suffix}.json"


def _load_result(results_dir, suffixes, algo):
    for suffix in suffixes:
        path = _result_path(results_dir, algo, suffix)
        if path.exists():
            return json.loads(path.read_text()), path
    return None, _result_path(results_dir, algo, suffixes[0])


def _setting_label(setting):
    if setting is None:
        return "no distortion"
    return str(setting).replace("-", "neg").replace(".", "p")


def _setting_display(dist_key, setting):
    if setting is None:
        return "—"
    if dist_key == "pitch_shift":
        return (f"+{setting}" if setting > 0 else str(setting)) + " ¢"
    if dist_key == "time_stretch":
        return f"{setting}×"
    if dist_key == "gaussian_noise":
        return f"{setting} dB"
    if dist_key == "bitcrush":
        return f"{setting} bit"
    if dist_key == "mp3_compression":
        return f"{setting} kbps"
    if dist_key == "background_noise":
        return str(setting)
    if dist_key == "cutting_audio":
        return f"{setting}%"
    if dist_key in {"high_pass_filter", "low_pass_filter"}:
        return f"{setting} Hz"
    if dist_key == "sample_suppression":
        return f"{setting}%"
    if dist_key == "resampling":
        return f"{setting} kHz"
    if dist_key == "reverberation":
        return str(setting)
    return str(setting)


def _metric_value(data, dist_key, setting, field_name):
    if data is None:
        return None
    if dist_key == "no_distortion":
        return data.get("no_distortion", {}).get(field_name)
    label = _setting_label(setting)
    return data.get(dist_key, {}).get(label, {}).get(field_name)


def _build_full_columns():
    columns = []
    for si, (section_name, dist_key, settings) in enumerate(SECTIONS):
        for setting in settings:
            columns.append((si, section_name, dist_key, setting, _setting_display(dist_key, setting)))
    return columns


def _status_fill(data):
    if data is None:
        return MISS_FILL
    meta = data.get("_meta", {})
    baseline = data.get("no_distortion", {})
    processed = meta.get("n_processed")
    total = meta.get("n_total")
    baseline_n = baseline.get("n")
    secs_n = baseline.get("n_secs")
    has_dist_errors = bool(meta.get("dist_errors"))
    if processed != total:
        return WARN_FILL
    if baseline_n is not None and processed is not None and baseline_n != processed:
        return WARN_FILL
    if secs_n is not None and processed is not None and secs_n != processed:
        return WARN_FILL
    if has_dist_errors:
        return WARN_FILL
    return OK_FILL


def _health_summary(data):
    if data is None:
        return "Missing result"
    meta = data.get("_meta", {})
    baseline = data.get("no_distortion", {})
    processed = meta.get("n_processed")
    total = meta.get("n_total")
    failed_embed = meta.get("n_failed_embed", 0)
    baseline_n = baseline.get("n")
    secs_n = baseline.get("n_secs")
    dist_errors = meta.get("dist_errors", {})
    notes = []
    if processed != total:
        notes.append(f"{processed}/{total} processed")
    if failed_embed:
        notes.append(f"{failed_embed} embed failures")
    if baseline_n is not None and processed is not None and baseline_n != processed:
        notes.append(f"baseline bit count {baseline_n}")
    if secs_n is not None and processed is not None and secs_n != processed:
        notes.append(f"baseline SECS count {secs_n}")
    if dist_errors:
        notes.append(f"{sum(dist_errors.values())} distortion errors across {len(dist_errors)} settings")
    if not notes:
        return "Clean result"
    return "; ".join(str(x) for x in notes)


def _affected_families(data):
    if data is None:
        return []
    meta = data.get("_meta", {})
    processed = meta.get("n_processed")
    if processed is None:
        return []
    families = set()
    for _section_name, dist_key, settings in SECTIONS:
        if dist_key == "no_distortion":
            continue
        for setting in settings:
            n = _metric_value(data, dist_key, setting, "n")
            n_secs = _metric_value(data, dist_key, setting, "n_secs")
            if (n is not None and n < processed) or (n_secs is not None and n_secs < processed):
                families.add(dist_key)
                break
    for key in meta.get("dist_errors", {}):
        family = key.split("/", 1)[0]
        families.add(family)
    return sorted(families)


def _build_metric_sheet(wb, dataset_key, results_dir, suffixes, algorithms, sheet_name, metric_field, number_format):
    ws = wb.create_sheet(sheet_name)
    columns = _build_full_columns()
    total_cols = 1 + len(columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title = ws.cell(1, 1)
    title.value = f"{dataset_key} / full robustness / {metric_field}"
    title.font = Font(bold=True, color="FFFFFF", size=13)
    title.fill = HDR_FILL
    title.alignment = CENTER

    ws.cell(2, 1, "Algorithm").font = Font(bold=True, color="FFFFFF")
    ws.cell(2, 1).fill = HDR_FILL
    ws.cell(2, 1).alignment = CENTER
    ws.cell(3, 1, "").fill = HDR_FILL

    start_col = 2
    idx = 0
    for si, (section_name, _dist_key, settings) in enumerate(SECTIONS):
        color = PatternFill("solid", fgColor=SECTION_COLORS[si % len(SECTION_COLORS)])
        span = len(settings)
        sc = start_col + idx
        ec = sc + span - 1
        if sc != ec:
            ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)
        cell = ws.cell(2, sc, section_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = color
        cell.alignment = CENTER
        idx += span

    for j, (si, _section_name, _dist_key, _setting, disp) in enumerate(columns, start=2):
        color = PatternFill("solid", fgColor=SECTION_COLORS[si % len(SECTION_COLORS)])
        cell = ws.cell(3, j, disp)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = color
        cell.alignment = CENTER

    row = 4
    for algo in algorithms:
        data, _path = _load_result(results_dir, suffixes, algo)
        row_fill = _status_fill(data)
        cell = ws.cell(row, 1, ALGO_LABELS.get(algo, algo))
        cell.font = Font(bold=True)
        cell.fill = row_fill
        cell.alignment = LEFT
        for j, (_si, _section_name, dist_key, setting, _disp) in enumerate(columns, start=2):
            value = _metric_value(data, dist_key, setting, metric_field)
            out = "-" if value is None else float(value)
            c = ws.cell(row, j, out)
            c.fill = row_fill
            c.alignment = CENTER
            if value is not None:
                c.number_format = number_format
        row += 1

    ws.column_dimensions["A"].width = 30
    for j, (_si, _section_name, dist_key, _setting, _disp) in enumerate(columns, start=2):
        width = 14 if dist_key in {"background_noise", "reverberation"} else 8
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 45
    ws.freeze_panes = "B4"
    return ws


def _build_health_sheet(wb, dataset_key, results_dir, suffixes, algorithms):
    ws = wb.create_sheet("Run Health")
    headers = [
        "Algorithm",
        "Status",
        "Result File",
        "Processed",
        "Total",
        "Failed Embed",
        "No Distortion Bit N",
        "No Distortion SECS N",
        "Distortion Errors",
        "Distortion Error Settings",
        "Decode Refusals",
        "Summary",
    ]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(1, col, text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = CENTER

    for algo in algorithms:
        data, path = _load_result(results_dir, suffixes, algo)
        row = ws.max_row + 1
        if data is None:
            values = [ALGO_LABELS.get(algo, algo), "MISSING", path.name, "", "", "", "", "", "", "", "", "Missing result"]
            fill = MISS_FILL
        else:
            meta = data.get("_meta", {})
            baseline = data.get("no_distortion", {})
            dist_errors = meta.get("dist_errors", {})
            fill = _status_fill(data)
            values = [
                ALGO_LABELS.get(algo, algo),
                "WARN" if fill == WARN_FILL else "OK",
                path.name,
                meta.get("n_processed"),
                meta.get("n_total"),
                meta.get("n_failed_embed"),
                baseline.get("n"),
                baseline.get("n_secs"),
                sum(dist_errors.values()),
                len(dist_errors),
                sum(meta.get("decode_refusals", {}).values()),
                _health_summary(data),
            ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.fill = fill
            cell.alignment = LEFT if col in {1, 3, 12} else CENTER
            if col == 1:
                cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    return ws


def _build_issue_focus_sheet(wb, results_dir, suffixes, algorithms):
    ws = wb.create_sheet("Validity Focus")
    headers = [
        "Algorithm",
        "Issue Type",
        "Affected Families",
        "Worst Bit N",
        "Worst SECS N",
        "Summary",
    ]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(1, col, text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = CENTER

    for algo in algorithms:
        data, _path = _load_result(results_dir, suffixes, algo)
        if data is None:
            continue
        meta = data.get("_meta", {})
        baseline = data.get("no_distortion", {})
        processed = meta.get("n_processed")
        dist_errors = meta.get("dist_errors", {})
        baseline_n = baseline.get("n")
        baseline_secs = baseline.get("n_secs")

        counts = []
        sec_counts = []
        for _section_name, dist_key, settings in SECTIONS:
            for setting in settings:
                n = _metric_value(data, dist_key, setting, "n")
                s = _metric_value(data, dist_key, setting, "n_secs")
                if n is not None:
                    counts.append(n)
                if s is not None:
                    sec_counts.append(s)
        worst_n = min(counts) if counts else None
        worst_secs = min(sec_counts) if sec_counts else None

        issues = []
        if processed is not None and meta.get("n_total") is not None and processed != meta.get("n_total"):
            issues.append(f"{processed}/{meta.get('n_total')} processed")
        if meta.get("n_failed_embed"):
            issues.append(f"{meta.get('n_failed_embed')} embed failures")
        if baseline_n is not None and processed is not None and baseline_n != processed:
            issues.append(f"baseline bit count {baseline_n}")
        if baseline_secs is not None and processed is not None and baseline_secs != processed:
            issues.append(f"baseline SECS count {baseline_secs}")
        if dist_errors:
            issues.append(f"{sum(dist_errors.values())} distortion errors across {len(dist_errors)} settings")
        if not issues:
            continue

        families = _affected_families(data)
        row = ws.max_row + 1
        values = [
            ALGO_LABELS.get(algo, algo),
            " / ".join(sorted(set(
                ["Embed failures"] if meta.get("n_failed_embed") else [] +
                ["Distortion errors"] if dist_errors else [] +
                ["Count mismatch"] if (
                    (baseline_n is not None and processed is not None and baseline_n != processed) or
                    (baseline_secs is not None and processed is not None and baseline_secs != processed)
                ) else []
            ))) or "Warn",
            ", ".join(families) if families else "Global",
            worst_n,
            worst_secs,
            "; ".join(issues),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.fill = WARN_FILL
            cell.alignment = LEFT if col in {1, 2, 3, 6} else CENTER
    ws.freeze_panes = "A2"
    return ws


def build_workbook(dataset_key, out_path, suffixes, algorithms):
    results_dir = PROJECT_DIR / "results" / "benchmark" / dataset_key
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets = [
        _build_metric_sheet(wb, dataset_key, results_dir, suffixes, algorithms, "Robustness Accuracy", "bit_accuracy", "0.0%"),
        _build_metric_sheet(wb, dataset_key, results_dir, suffixes, algorithms, "Robustness SECS", "secs", "0.000"),
        _build_health_sheet(wb, dataset_key, results_dir, suffixes, algorithms),
        _build_issue_focus_sheet(wb, results_dir, suffixes, algorithms),
    ]
    for ws in sheets:
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row_cells in ws.iter_rows(min_col=col, max_col=col):
                value = row_cells[0].value
                max_len = max(max_len, len(str(value)) if value is not None else 0)
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)
        if ws.title.startswith("Robustness"):
            ws.freeze_panes = "B4"
    wb.save(out_path)


def main():
    args = parse_args()
    suffixes = [s for s in args.suffixes.split(",") if s]
    algorithms = [s for s in args.algorithms.split(",") if s] or ALGORITHMS
    unknown = [algo for algo in algorithms if algo not in ALGORITHMS]
    if unknown:
        raise SystemExit(f"Unknown algorithms: {', '.join(unknown)}")
    out_path = PROJECT_DIR / args.out
    build_workbook(args.dataset, out_path, suffixes, algorithms)
    print(f"Saved -> {out_path}")


if __name__ == "__main__":
    main()
