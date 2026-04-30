"""
15_visqol_unwm.py  —  ViSQOL scores for unwatermarked distorted audio.

Applies all 116 distortions to clean original audio (no watermarking) and
computes ViSQOL of (distorted original vs clean original).  Only ViSQOL is
computed — no embedding, decoding, PESQ, or ESTOI.

Datasets:
  speech_ljspeech  (speech mode, 16 kHz)
  music_m4singer   (audio  mode, 48 kHz)
  speech_daps      (speech mode, 16 kHz)
  music_moisesdb   (audio  mode, 48 kHz)

By default the script evaluates the full dataset for each target. Set
`SOK_N_SAMPLES=<int>` to run a smaller reproducibility subset.

Run:
  nohup envs/viz/bin/python -u scripts/15_visqol_unwm.py \\
        > logs/15_visqol_unwm.log 2>&1 & disown $!

Output:
  results/visqol_unwm/{dataset_key}.json    per-distortion mean ViSQOL
  results/visqol_unwm/{dataset_key}.checkpoint.json  (resume support)
  results/robustness_accuracy.xlsx          new sheet "Unwm ViSQOL" appended
"""
# ── Thread limits before any import touching numpy/scipy ─────────────────────
import os
os.environ.setdefault("NUMBA_THREADING_LAYER", "omp")
os.environ.setdefault("OMP_NUM_THREADS",       "1")
os.environ.setdefault("NUMBA_NUM_THREADS",     "1")
os.environ.setdefault("OPENBLAS_NUM_THREADS",  "1")
os.environ.setdefault("MKL_NUM_THREADS",       "1")

import json
import sys
import numpy as np
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

PROJECT_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).parent))

import benchmark_utils as bu
from config import DATASET_DIR as CONFIG_DATASET_DIR

# ── Config ────────────────────────────────────────────────────────────────────
DATASET_DIR  = Path(os.environ.get("SOK_DATASET_ROOT", str(CONFIG_DATASET_DIR)))
OUT_DIR      = PROJECT_DIR / "results" / "visqol_unwm"
XLSX_PATH    = PROJECT_DIR / "results" / "robustness_accuracy.xlsx"
LOG_INTERVAL = 10
N_WORKERS    = min(8, os.cpu_count() or 4)
SEED         = 42
_N_SAMPLES_RAW = os.environ.get("SOK_N_SAMPLES", "").strip()
DEFAULT_N_SAMPLES = None if not _N_SAMPLES_RAW else int(_N_SAMPLES_RAW)

TARGETS = [
    # (dataset_key, category, folder, n_samples, visqol_mode)
    ("speech_ljspeech", "speech", "LJSpeech-1.1", DEFAULT_N_SAMPLES, "speech"),
    ("music_m4singer",  "music",  "m4singer",      DEFAULT_N_SAMPLES, "audio"),
    ("speech_daps",     "speech", "daps",          DEFAULT_N_SAMPLES, "speech"),
    ("music_moisesdb",  "music",  "moisesdb",      DEFAULT_N_SAMPLES, "audio"),
]

DATASET_LABELS = {
    "speech_ljspeech": "Speech-LJSpeech  (speech mode)",
    "music_m4singer":  "Music-M4Singer   (audio mode)",
    "speech_daps":     "Speech-DAPS      (speech mode)",
    "music_moisesdb":  "Music-MoisesDB   (audio mode)",
}

# ── Excel style constants (mirrors 12_excel_report.py) ───────────────────────
_SECTION_DEFS = [
    ("No Distortion",          "no_distortion"),
    ("Pitch Shift (cents)",    "pitch_shift"),
    ("Time Stretch",           "time_stretch"),
    ("Gaussian Noise (SNR)",   "gaussian_noise"),
    ("Bitcrush (bits)",        "bitcrush"),
    ("MP3 Compression",        "mp3_compression"),
    ("Background Noise (SNR)", "background_noise"),
    ("Cutting Audio",          "cutting_audio"),
    ("High-Pass Filter (Hz)",  "high_pass_filter"),
    ("Low-Pass Filter (Hz)",   "low_pass_filter"),
    ("Sample Suppression",     "sample_suppression"),
    ("Resampling (kHz)",       "resampling"),
    ("Reverberation",          "reverberation"),
]

SECTION_COLORS = [
    "4472C4", "70AD47", "ED7D31", "9E480E", "A5A5A5",
    "FFC000", "5B9BD5", "C00000", "00B0F0", "7030A0",
    "FF0000", "92D050", "00B050",
]
ROW_FILLS    = ["DEEAF1", "E2EFDA"]
HDR_FILL     = "2E4057"
VISQOL_FILL  = "F4CCFF"


# ── Display helpers ───────────────────────────────────────────────────────────
def _setting_display(dist_key, s):
    if s is None:
        return "—"
    if dist_key == "pitch_shift":
        return (f"+{s}" if s > 0 else str(s)) + " ¢"
    if dist_key == "time_stretch":
        return f"{s}×"
    if dist_key == "gaussian_noise":
        return f"{s} dB"
    if dist_key == "bitcrush":
        return f"{s} bit"
    if dist_key == "mp3_compression":
        return f"{s} kbps"
    if dist_key == "background_noise":
        return s
    if dist_key == "cutting_audio":
        return f"{s}%"
    if dist_key in ("high_pass_filter", "low_pass_filter"):
        return f"{s} Hz"
    if dist_key == "sample_suppression":
        return f"{s}%"
    if dist_key == "resampling":
        return f"{s} kHz"
    if dist_key == "reverberation":
        return s
    return str(s)


def _build_columns():
    """Ordered list of (section_idx, sec_name, dist_key, setting, display_label)."""
    cols = []
    for si, (sec_name, dist_key) in enumerate(_SECTION_DEFS):
        settings = [None] if dist_key == "no_distortion" else bu.DISTORTIONS[dist_key]
        for s in settings:
            cols.append((si, sec_name, dist_key, s, _setting_display(dist_key, s)))
    return cols


# ── File sampling (same seed=42 as large_scale_utils) ────────────────────────
def _sample_files(category, folder, n):
    dataset_path = DATASET_DIR / category / folder
    all_wavs = sorted([p for p in dataset_path.rglob("*.wav")
                       if not p.name.startswith("._")])
    if not all_wavs:
        raise FileNotFoundError(f"No WAV files in {dataset_path}")
    if n is None or n >= len(all_wavs):
        return all_wavs
    rng = np.random.RandomState(SEED)
    idx = sorted(rng.choice(len(all_wavs), size=n, replace=False))
    return [all_wavs[i] for i in idx]


# ── Per-dataset computation ───────────────────────────────────────────────────
def _run_dataset(dataset_key, category, folder, n_samples, mode):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    json_path = OUT_DIR / f"{dataset_key}.json"
    ckpt_path = OUT_DIR / f"{dataset_key}.checkpoint.json"

    if json_path.exists():
        print(f"\n[SKIP] {dataset_key}: result JSON already exists", flush=True)
        return

    print(f"\n  === {dataset_key}  ({n_samples} files, visqol_mode={mode}) ===",
          flush=True)

    # Build distortion key list: (acc_key, dist_name, setting, label)
    dist_keys = []
    for dist_name, settings in bu.DISTORTIONS.items():
        for s in settings:
            label = bu.setting_label(s)
            dist_keys.append((f"{dist_name}/{label}", dist_name, s, label))

    # Accumulator: flat dict of {key: {"s": sum, "n": count}}
    def _new_acc():
        d = {"no_distortion": {"s": 0.0, "n": 0}}
        for k, *_ in dist_keys:
            d[k] = {"s": 0.0, "n": 0}
        return d

    acc = _new_acc()
    files_done = set()
    n_processed = 0
    n_failed = 0

    # Resume from checkpoint
    if ckpt_path.exists():
        try:
            ckpt = json.loads(ckpt_path.read_text())
            acc = ckpt["acc"]
            files_done = set(ckpt["files_done"])
            n_processed = ckpt["n_processed"]
            print(f"  Resumed from checkpoint: {n_processed} files done", flush=True)
        except Exception as e:
            print(f"  Checkpoint load failed ({e}), starting fresh", flush=True)
            acc = _new_acc()
            files_done = set()
            n_processed = 0

    all_files = _sample_files(category, folder, n_samples)
    todo = [f for f in all_files if f.stem not in files_done]
    print(f"  {len(all_files)} sampled, {len(todo)} remaining", flush=True)

    def _accum(key, v):
        if v == v:  # not NaN
            acc[key]["s"] += v
            acc[key]["n"] += 1

    def _save_ckpt():
        tmp = ckpt_path.with_suffix(".tmp")
        tmp.write_text(json.dumps({
            "n_processed": n_processed,
            "files_done":  list(files_done),
            "acc":         acc,
        }))
        tmp.replace(ckpt_path)

    with ThreadPoolExecutor(max_workers=N_WORKERS) as executor:

        def _safe(fut, timeout=300):
            try:
                return float(fut.result(timeout=timeout))
            except Exception:
                return float("nan")

        for path in todo:
            stem = path.stem
            try:
                orig, sr = bu.load_audio_clipped(str(path), stem=stem)
            except Exception as e:
                print(f"  LOAD SKIP {stem}: {e}", flush=True)
                n_failed += 1
                continue

            # Baseline: undistorted (orig vs orig, expect ~5.0)
            bsn_fut = executor.submit(
                bu.compute_visqol, orig.copy(), orig.copy(), sr, mode)

            # Apply each distortion; submit ViSQOL to thread pool
            vfuts = []  # [(acc_key, future_or_None)]
            for key, dist_name, setting, label in dist_keys:
                try:
                    dist = bu.apply_distortion(orig, sr, dist_name, setting, stem=stem)
                    n    = min(len(dist), len(orig))
                    vfuts.append((key, executor.submit(
                        bu.compute_visqol,
                        orig[:n].copy(), dist[:n].copy(), sr, mode)))
                except Exception:
                    vfuts.append((key, None))

            # Collect (pool has been running during the distortion loop above)
            _accum("no_distortion", _safe(bsn_fut))
            for key, fut in vfuts:
                _accum(key, _safe(fut) if fut is not None else float("nan"))

            n_processed += 1
            files_done.add(stem)

            if n_processed == 1 or n_processed % LOG_INTERVAL == 0:
                print(f"    [{n_processed}/{len(all_files)}] {stem}", flush=True)
            if n_processed % LOG_INTERVAL == 0:
                _save_ckpt()

    _save_ckpt()

    # Write final JSON (same structure as existing benchmark JSONs, visqol only)
    def _mean(key):
        d = acc[key]
        return float(d["s"] / d["n"]) if d["n"] > 0 else float("nan")

    result = {
        "_meta": {
            "dataset":     dataset_key,
            "n_processed": n_processed,
            "n_failed":    n_failed,
            "mode":        mode,
        },
        "no_distortion": {"visqol": _mean("no_distortion")},
    }
    for dist_name, settings in bu.DISTORTIONS.items():
        result[dist_name] = {}
        for s in settings:
            label = bu.setting_label(s)
            result[dist_name][label] = {"visqol": _mean(f"{dist_name}/{label}")}

    json_path.write_text(json.dumps(result, indent=2))
    print(f"  → {json_path}  ({n_processed} processed, {n_failed} failed)",
          flush=True)


# ── Excel sheet builder ───────────────────────────────────────────────────────
def _add_unwm_sheet(wb, results):
    """
    Append sheet 'Unwm ViSQOL' to wb (does NOT touch existing sheets).
    results: OrderedDict {dataset_key: json_dict}.

    Layout:
      Row 1  : section header spans (colored, merged)
      Row 2  : distortion setting labels
      Rows 3+: one row per dataset, ViSQOL value per distortion condition
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    def _fill(hex_c):
        return PatternFill("solid", fgColor=hex_c)

    def _hfont(color="FFFFFF", bold=True):
        return Font(bold=bold, color=color)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    ws          = wb.create_sheet("Unwm ViSQOL")
    columns     = _build_columns()
    COL_DATASET = 1
    COL_START   = 2

    # ── Row 1: section header spans ──
    ws.cell(1, COL_DATASET).fill = _fill(HDR_FILL)

    section_spans = {}
    for j, (si, *_) in enumerate(columns):
        col = COL_START + j
        section_spans.setdefault(si, [col, col])
        section_spans[si][1] = col

    for si, (sc, ec) in section_spans.items():
        sec_name = _SECTION_DEFS[si][0]
        color    = SECTION_COLORS[si % len(SECTION_COLORS)]
        if sc != ec:
            ws.merge_cells(start_row=1, start_column=sc,
                           end_row=1,   end_column=ec)
        c = ws.cell(1, sc, sec_name)
        c.font = _hfont(); c.fill = _fill(color); c.alignment = center

    # ── Row 2: setting labels ──
    c = ws.cell(2, COL_DATASET, "Dataset")
    c.font = _hfont(); c.fill = _fill(HDR_FILL); c.alignment = center

    for j, (si, sec_name, dist_key, s, disp) in enumerate(columns):
        color = SECTION_COLORS[si % len(SECTION_COLORS)]
        c = ws.cell(2, COL_START + j, disp)
        c.font = _hfont(); c.fill = _fill(color); c.alignment = center

    # ── Data rows: one per dataset ──
    for ri, (dataset_key, json_data) in enumerate(results.items()):
        row      = 3 + ri
        row_fill = _fill(ROW_FILLS[ri % 2])
        label    = DATASET_LABELS.get(dataset_key, dataset_key)

        c = ws.cell(row, COL_DATASET, label)
        c.font = Font(bold=True); c.fill = row_fill; c.alignment = left

        for j, (si, sec_name, dist_key, s, disp) in enumerate(columns):
            if dist_key == "no_distortion":
                v = json_data.get("no_distortion", {}).get("visqol", float("nan"))
            else:
                lbl = bu.setting_label(s)
                v   = json_data.get(dist_key, {}).get(lbl, {}).get("visqol", float("nan"))
            try:
                v   = float(v)
                txt = f"{v:.3f}" if v == v else "-"
            except (TypeError, ValueError):
                txt = "-"
            c = ws.cell(row, COL_START + j, txt)
            c.fill = _fill(VISQOL_FILL); c.alignment = center

    # ── Column widths / row heights ──
    ws.column_dimensions[get_column_letter(COL_DATASET)].width = 40
    for j, (si, sec_name, dist_key, *_) in enumerate(columns):
        w = 14 if dist_key in ("background_noise", "reverberation") else 8
        ws.column_dimensions[get_column_letter(COL_START + j)].width = w
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 45
    for ri in range(len(results)):
        ws.row_dimensions[3 + ri].height = 22
    ws.freeze_panes = "B3"


def _update_excel():
    import openpyxl

    if XLSX_PATH.exists():
        wb = openpyxl.load_workbook(str(XLSX_PATH))
        print(f"Opened existing: {XLSX_PATH}", flush=True)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        print(f"Creating new workbook: {XLSX_PATH}", flush=True)

    # Idempotent: remove old sheet if re-running
    if "Unwm ViSQOL" in wb.sheetnames:
        del wb["Unwm ViSQOL"]
        print("  Removed old 'Unwm ViSQOL' sheet", flush=True)

    results = {}
    for dataset_key, *_ in TARGETS:
        p = OUT_DIR / f"{dataset_key}.json"
        if p.exists():
            results[dataset_key] = json.loads(p.read_text())
            print(f"  Loaded {dataset_key}", flush=True)
        else:
            print(f"  WARNING: {p.name} not found — skipped", flush=True)

    if not results:
        print("No result JSONs found; skipping Excel update.", flush=True)
        return

    _add_unwm_sheet(wb, results)
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(XLSX_PATH))
    print(f"Saved → {XLSX_PATH}  (sheets: {wb.sheetnames})", flush=True)


# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    print(f"N_WORKERS={N_WORKERS}  SEED={SEED}", flush=True)
    print(f"OUT_DIR   = {OUT_DIR}", flush=True)
    print(f"XLSX_PATH = {XLSX_PATH}", flush=True)

    for dataset_key, category, folder, n_samples, mode in TARGETS:
        _run_dataset(dataset_key, category, folder, n_samples, mode)

    print("\n── Updating Excel ──────────────────────────────────────────────",
          flush=True)
    _update_excel()
    print("Done.", flush=True)


if __name__ == "__main__":
    main()
